#!/usr/bin/env python3
"""
scripts/python/db_diag_report.py
Read *.tsv files from --indir and write a single Excel workbook (--out),
one worksheet per TSV. Requires 'openpyxl'. If not installed, exits non-zero
so the caller can print a friendly message (TSVs still exist).
"""
import argparse
import os
import glob

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--indir", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    try:
        from openpyxl import Workbook
    except Exception as e:
        print(f"[db_diag_report] openpyxl missing or failed to import: {e}")
        return 1

    tsv_paths = sorted(glob.glob(os.path.join(args.indir, "*.tsv")))
    if not tsv_paths:
        print("[db_diag_report] No TSV files found; nothing to write.")
        return 1

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "index"
    ws0.append(["sheet_name", "source_file"])

    for p in tsv_paths:
        sheet_name = os.path.splitext(os.path.basename(p))[0]
        # Excel sheet name constraints
        safe = sheet_name[:31]
        safe = safe.replace("/", "_").replace("\\", "_").replace("*", "_").replace("?", "_").replace("[", "(").replace("]", ")")
        ws = wb.create_sheet(title=safe)

        with open(p, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                # TSV -> list
                row = line.rstrip("\n").split("\t")
                ws.append(row)

        ws0.append([safe, os.path.basename(p)])

    # Remove the default first empty sheet if nothing written there
    if ws0.max_row == 1 and ws0.max_column == 1:
        wb.remove(ws0)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)
    print(f"[db_diag_report] Wrote Excel: {args.out}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
